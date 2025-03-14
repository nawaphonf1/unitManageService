from sqlalchemy.orm import Session, class_mapper
from sqlalchemy.sql import func, and_, or_
from sqlalchemy import desc

from app.models.unit import Unit
from app.models.position import Position
from app.models.dept import Dept
from app.models.mission import Mission
from app.models.mission_unit import MissionUnit

from app.schemas.unit import UnitCreate, UnitAll, UnitUpdate
from app.schemas.mission import UpdateMissionUnitParam

from app.services import unit_service

from datetime import date
from io import BytesIO
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class MissionService:
    def del_mission(db, mission_id):
        print(mission_id)
        # อัปเดตค่า is_active เป็น 0 โดยระบุเงื่อนไขที่ต้องการ
        db.query(Mission).filter(Mission.mission_id == mission_id).update({"is_active": 0})
        db.commit()  # อย่าลืม commit เพื่อบันทึกการเปลี่ยนแปลง

    def create_mission(db: Session, data: UpdateMissionUnitParam):
        mission = Mission()

        mission.mission_name = data.mission_name
        mission.mission_start = data.mission_start
        mission.mission_end = data.mission_end
        mission.mission_detail = data.mission_detail
        mission.mission_type = data.mission_type
        mission.mission_status = "r"  # ใช้สถานะที่มีความหมาย เช่น "ready" แทน "r"
        
        db.add(mission)
        db.commit()  # ทำการ commit เพื่อบันทึกภารกิจ
        db.refresh(mission)  # รีเฟรช mission หลังจาก commit เพื่อให้ได้ข้อมูลที่อัปเดตล่าสุด

        # เพิ่ม mission units ใหม่
        for unit_id in data.mission_unit_id:
            new_mission_unit = MissionUnit(mission_id=mission.mission_id, unit_id=unit_id)
            db.add(new_mission_unit)

        db.commit()  # ทำการ commit เพื่อบันทึก mission units ใหม่

        return mission


    def get_mission_by_units_id(db: Session, unit_id: int):
        query = db.query(
            Mission.mission_id,
            MissionUnit.unit_id,
            Mission.mission_name,
            Mission.mission_start,
            Mission.mission_end,
            Mission.mission_status,
        ).\
        join(MissionUnit, Mission.mission_id == MissionUnit.mission_id).\
        filter(and_(
            MissionUnit.unit_id == unit_id,
            Mission.is_active == True
        )).all()

        return query
    
    def get_missions(db: Session, skip: int = 0, limit: int = 100, mission_name: str = None, mission_start: date = None, mission_end: date = None, mission_type: str = None, mission_status: str = None):
        query = db.query(
            Mission.mission_id,
            Mission.mission_name,
            Mission.mission_start,
            Mission.mission_end,
            Mission.mission_status,
            Mission.is_active,
            Mission.created_at,
            Mission.mission_type
        ).filter(Mission.is_active == True)
        
        
        if mission_name:
            query = query.filter(Mission.mission_name.contains(mission_name))
        if mission_start:
            query = query.filter(Mission.mission_start == mission_start)
        if mission_end:
            query = query.filter(Mission.mission_end == mission_end)
        if mission_type:
            query = query.filter(Mission.mission_type == mission_type)
        if mission_status:
            query = query.filter(Mission.mission_status == mission_status)
        # Get total count for pagination
        total = query.with_entities(func.count()).scalar()

        # Apply pagination
        missions = query.order_by(desc(Mission.created_at)).offset(skip).limit(limit).all()
        
        return {
        "missions": [
            {
                "mission_id": mission.mission_id,
                "mission_name": mission.mission_name,
                "mission_start": mission.mission_start,
                "mission_end": mission.mission_end,
                "mission_status": mission.mission_status,
                "is_active": mission.is_active,
                "created_at": mission.created_at,
                "mission_type": mission.mission_type
            }
            for mission in missions
        ],
        "total": total,
        "skip": skip,
        "limit": limit,
    }

    def get_mission_by_id(db: Session, mission_id: int):
        query = db.query(
            Mission.mission_id,
            Mission.mission_name,
            Mission.mission_start,
            Mission.mission_end,
            Mission.mission_status,
            Mission.is_active,
            Mission.created_at,
            Mission.mission_type,
            Mission.mission_detail
        ).filter(and_(
            Mission.mission_id == mission_id,
            Mission.is_active == True
        )).first()

        if query is None:
            return None  # คืน None เพื่อให้ตรวจสอบได้ใน Controller
        

        
        # แปลงผลลัพธ์ให้เป็น dictionary
        return {
            "mission_id": query.mission_id,
            "mission_name": query.mission_name,
            "mission_start": query.mission_start,
            "mission_end": query.mission_end,
            "mission_status": query.mission_status,
            "is_active": query.is_active,
            "created_at": query.created_at,
            "mission_type": query.mission_type,
            "mission_detail": query.mission_detail
        }

    def get_unit_by_mission_id(db: Session, mission_id: int):
        query = db.query(
                    MissionUnit.mission_id,
                    Unit.units_id,
                    Unit.first_name,
                    Unit.last_name,
                    Position.position_id,
                    Position.position_name,
                    Position.position_name_short,
                    Position.position_seq,
                ).\
                join(Unit,Unit.units_id == MissionUnit.unit_id).\
                join(Position, Position.position_id == Unit.position_id).\
                filter(MissionUnit.mission_id ==mission_id).\
                order_by(desc(Position.position_seq)).\
                all()
        
        return query

    def update_mission_units_service(db, mission_id: int, data: UpdateMissionUnitParam):
        # ดึง mission ที่ต้องการแก้ไขจากฐานข้อมูล
        mission = db.query(Mission).filter(Mission.mission_id == mission_id).first()

        if not mission:
            return None  # ถ้าไม่เจอ mission คืนค่า None เพื่อแจ้งว่าไม่พบข้อมูล

        # อัปเดตข้อมูล mission
        mission.mission_name = data.mission_name
        mission.mission_start = data.mission_start
        mission.mission_end = data.mission_end
        mission.mission_detail = data.mission_detail
        mission.mission_type = data.mission_type
        mission.mission_status = data.mission_status

        # ลบ mission units เก่าที่เกี่ยวข้อง
        db.query(MissionUnit).filter(MissionUnit.mission_id == mission_id).delete()

        # เพิ่ม mission units ใหม่
        for unit_id in data.mission_unit_id:
            new_mission_unit = MissionUnit(mission_id=mission_id, unit_id=unit_id)
            db.add(new_mission_unit)

        db.commit()  # บันทึกการเปลี่ยนแปลงลงฐานข้อมูล
        db.refresh(mission)  # รีเฟรชข้อมูล mission หลังการแก้ไข

        return mission


    def update_status_unit(db: Session):
        MissionService.update_status_mission(db)

        current_date = date.today()

        # ดึง units ที่ไม่พร้อมใช้งาน
        units = db.query(Unit).\
            join(MissionUnit, MissionUnit.unit_id == Unit.units_id).\
            join(Mission, Mission.mission_id == MissionUnit.mission_id).\
            filter(and_(
                Mission.mission_start <= current_date,
                Mission.mission_end >= current_date,
                Mission.is_active == True
            )).all()

        # สร้างรายการ units_id ที่ไม่พร้อม
        unit_is_not_ready = {unit.units_id for unit in units}

        # ดึง units ทั้งหมด
        all_units = db.query(Unit).filter(Unit.is_active == True).all()
        # อัปเดตสถานะในหน่วยความจำ
        for unit in all_units:
            unit.status = "not ready" if unit.units_id in unit_is_not_ready else "ready"

        # บันทึกการเปลี่ยนแปลงครั้งเดียว
        db.commit()

        

    def update_status_mission(db:Session):
        current_date = date.today()
        mission =  db.query(Mission).filter(Mission.is_active == True).all()
        for mission in mission:
            if mission.mission_start <=  current_date and mission.mission_end >= current_date:
                mission.mission_status = "r" 
            elif mission.mission_start > current_date:
                mission.mission_status = "w"
            else:
                mission.mission_status = "n"
        
        db.commit()


    # ฟังก์ชันช่วยแปลงค่า NaN และ float ให้เป็น string
    def safe_str(value):
        if pd.isna(value):  # ถ้าเป็น NaN ให้คืนค่าเป็น ""
            return ""
        return str(int(value)) if isinstance(value, float) else str(value)



    async def import_excel(db, file):
        try:
            # อ่านไฟล์ Excel จาก memory (BytesIO) ด้วย pandas
            content = await file.read()
            df = pd.read_excel(BytesIO(content))
            # ตรวจสอบข้อมูลใน Excel
            for index, row in df.iterrows():
                dept_id = None
                # # ตรวจสอบว่าค่าของ 'Unnamed: 1' หรือคอลัมน์อื่น ๆ ไม่เป็น NaN
                if pd.notna(row['ชื่อ']) and pd.notna(row['นามสกุล']):
                    position = db.query(Position.position_id,Position.position_name_short).filter(Position.position_name_short.contains(row['ยศ'])).first()
                    if(position):
                        position_id = position.position_id
                        # ตรวจสอบและลบช่องว่างจากคอลัมน์ 2, 3, 4, 16 ถึง 20
                        first_name = row['ชื่อ']
                        last_name = row['นามสกุล']      
                                         
                        tel = MissionService.safe_str(row['เบอร์โทร'])
                        identify_id = MissionService.safe_str(row['หมายเลขประจำตัว'])
                        blood_group_id = MissionService.safe_str(row['กรุ๊ปเลือด'])
                        address_detail = MissionService.safe_str(row['ที่อยู่'])
                        identify_soldier_id = MissionService.safe_str(row['หมายเลขประจำตัวทหาร'])
                        dept = MissionService.safe_str(row['สังกัด'])
                        position_detail = MissionService.safe_str(row['ตำแหน่ง'])
                        
                        if pd.notna(row['สังกัด']):
                            dept_data = db.query(Dept.dept_id,Dept.dept_name_short).filter(Dept.dept_name_short == dept).first()
                            dept_id = dept_data.dept_id
        
                        else:
                            dept_id = None

                        print(identify_soldier_id)
                        duplicate_unit = db.query(Unit).filter(Unit.identify_soldier_id == str(identify_soldier_id)).first()
                        if(duplicate_unit):                            
                            updateUnit = UnitUpdate(
                                first_name = first_name,
                                last_name = last_name,
                                position_id = position_id,
                                identify_soldier_id = identify_soldier_id,
                                identify_id = identify_id,
                                tel = tel,
                                blood_group_id = blood_group_id,
                                address_detail = address_detail,
                                dept_id = dept_id,
                                position_detail = position_detail
                            )
                            unit_service.update_unit(db,duplicate_unit.units_id ,updateUnit)

                        else:
                            #สร้างข้อมูล
                            dataCeateUnit = UnitCreate(
                                first_name = first_name,
                                last_name = last_name,
                                position_id = position_id,
                                identify_soldier_id = identify_soldier_id,
                                identify_id = identify_id,
                                tel = tel,
                                blood_group_id = blood_group_id,
                                address_detail = address_detail,
                                dept_id = dept_id,
                                position_detail = position_detail

                            )
                            unit_service.create_unit(db, dataCeateUnit)


        except Exception as e:
            print(e)
            return {"error": str(e)}

    

    def export_mission(db, mission_id):
        mission_data = db.query(
                    Mission.mission_id,
                    Mission.mission_name,
                    Mission.mission_start,
                    Mission.mission_end,
                    Mission.mission_status,
                    Mission.mission_detail
                ).filter(Mission.mission_id == mission_id).first()
        
        mission_unit_data = db.query(
                    MissionUnit.mission_id,
                    Unit.units_id,
                    Unit.first_name,
                    Unit.last_name,
                    Unit.position_detail,
                    Position.position_name,
                    Position.position_name_short,
                    Position.position_seq,
                    Dept.dept_name_short,
                    Unit.identify_id,
        ).\
        outerjoin(Unit,Unit.units_id == MissionUnit.unit_id).\
        outerjoin(Position, Position.position_id == Unit.position_id).\
        outerjoin(Dept, Dept.dept_id == Unit.dept_id).\
        filter(MissionUnit.mission_id == mission_id).\
        order_by(desc(Position.position_seq)).\
        all()

        # สร้าง DataFrame สำหรับตารางข้อมูล
        orders_list = []
        for index, item in enumerate(mission_unit_data, start=1):
            orders_list.append({
                "ลำดับ": index,
                "ยศ": item.position_name_short,
                "ชื่อ": item.first_name,
                "นามสกุล": item.last_name,
                "สังกัด": item.dept_name_short,
                "ตำแหน่ง":item.position_detail
            })
        df = pd.DataFrame(orders_list)

        # สร้างไฟล์ Excel
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=4, sheet_name="รายงาน")  # ข้อมูลเริ่มที่แถวที่ 5
            workbook = writer.book
            worksheet = writer.sheets["รายงาน"]

            mission_status = None
            if mission_data.mission_status == "r":
                mission_status = "กำลังดำเนินการ"
            elif mission_data.mission_status == "w":
                mission_status = "รอดำเนินการ"
            else:
                mission_status = "สิ้นสุดแล้ว"
            # เพิ่ม Header Mission
            worksheet["A1"] = "ชื่อภารกิจ:"
            worksheet["B1"] = mission_data.mission_name
            worksheet["A2"] = "วันที่เริ่มต้น:"
            worksheet["B2"] = MissionService.convert_to_thai_date(mission_data.mission_start)
            worksheet["C2"] = "วันที่สิ้นสุด:"
            worksheet["D2"] = MissionService.convert_to_thai_date(mission_data.mission_end)
            worksheet["A3"] = "รายละเอียดภารกิจ:"
            worksheet["B3"] = mission_data.mission_detail
            worksheet["A4"] = "สถานะ:"
            worksheet["B4"] = mission_status

            # Merge Cell สำหรับหัวข้อ Mission
            worksheet.merge_cells("B1:D1")
            worksheet.merge_cells("B3:D3")

            # ปรับความกว้างของคอลัมน์
            for col in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max_length + 2

            # กำหนด Style สำหรับ Header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            alignment = Alignment(horizontal="center", vertical="center")

            # จัดการหัวตาราง
            for cell in worksheet[5]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border_style
                cell.alignment = alignment

            # ปรับ Style สำหรับข้อมูลในตาราง
            for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row, min_col=1, max_col=6):
                for cell in row:
                    cell.border = border_style
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # ส่งไฟล์ Excel
        excel_file.seek(0)
        return excel_file

    def convert_to_thai_date(date):
        months_thai = [
            "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", 
            "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", 
            "พฤศจิกายน", "ธันวาคม"
        ]
        # แปลงเดือนและปี
        month_thai = months_thai[date.month - 1]
        year_thai = date.year + 543  # แปลงปี ค.ศ. เป็น พ.ศ.
        return f"{date.day} {month_thai} {year_thai}"

    def export_mission_unit(db, mission_name, mission_start, mission_end, mission_type, mission_status, position_detail, position_name):
        # ดึงข้อมูลภารกิจที่อยู่ในช่วงวันที่กำหนด
        mission_data = db.query(
            Mission.mission_id,
            MissionUnit.unit_id,
            Mission.mission_name,
            Mission.mission_start,
            Mission.mission_end
        ).join(MissionUnit, Mission.mission_id == MissionUnit.mission_id)

        if mission_start and mission_end:
            mission_data = mission_data.filter(
                and_(
                    Mission.mission_start >= mission_start,
                    Mission.mission_end <= mission_end
                )
            )
        if mission_name:
            mission_data = mission_data.filter(Mission.mission_name.contains(mission_name))
        if mission_type:
            mission_data = mission_data.filter(Mission.mission_type == mission_type)
        if mission_status:
            mission_data = mission_data.filter(Mission.mission_status == mission_status)

        mission_data = mission_data.all()
        # ดึงข้อมูลหน่วย
        unit_data = db.query(
            Unit.units_id,
            Position.position_name_short,
            Unit.first_name,
            Unit.last_name,
            Unit.position_detail
        ).join(Position, Position.position_id == Unit.position_id).\
        order_by(desc(Position.position_seq)).\
        filter(Unit.is_active == True)

        if position_detail:
            unit_data = unit_data.filter(Unit.position_detail.contains(position_detail))
        if position_name:
            unit_data = unit_data.filter(or_(
                Position.position_name.contains(position_name),
                Position.position_name_short.contains(position_name)
            ))

        unit_data = unit_data.all()
        # แปลงชื่อภารกิจให้มีวันที่แบบ พ.ศ.
        mission_names = list(set(
            f"{m.mission_name} ({MissionService.convert_to_thai_date(m.mission_start)} - {MissionService.convert_to_thai_date(m.mission_end)})"
            for m in mission_data
        ))

        # สร้าง DataFrame
        orders_list = []
        for index, unit in enumerate(unit_data, start=1):
            unit_missions = {mission_name: "" for mission_name in mission_names}  # ตั้งค่าเริ่มต้นเป็น ""

            for mission in mission_data:
                mission_full_name = f"{mission.mission_name} ({MissionService.convert_to_thai_date(mission.mission_start)} - {MissionService.convert_to_thai_date(mission.mission_end)})"
                if mission.unit_id == unit.units_id:
                    unit_missions[mission_full_name] = "/"  # ถ้าตรงกันให้เป็น "/"

            # รวมข้อมูลหน่วยและภารกิจ
            unit_entry = {
                "ลำดับ": index,
                "ยศ": unit.position_name_short,
                "ชื่อ": unit.first_name,
                "นามสกุล": unit.last_name,
                "ตำแหน่ง": unit.position_detail,
                **unit_missions  # รวมข้อมูลภารกิจเข้าไปใน dictionary
            }
            orders_list.append(unit_entry)

        # แปลงเป็น DataFrame
        df = pd.DataFrame(orders_list)

        # สร้างไฟล์ Excel
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=1, sheet_name="รายงาน")  # ข้อมูลเริ่มที่แถวที่ 5
            workbook = writer.book
            worksheet = writer.sheets["รายงาน"]

            # ปรับความกว้างของคอลัมน์
            for col in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max_length + 2

            # กำหนด Style สำหรับ Header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            alignment = Alignment(horizontal="center", vertical="center")

            # จัดการหัวตาราง
            for cell in worksheet[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border_style
                cell.alignment = alignment

            # **หมุนข้อความของ Header ตั้งแต่คอลัมน์ที่ F เป็นต้นไป**
            for col_idx in range(6, worksheet.max_column + 1):  # เริ่มที่คอลัมน์ที่ 6 (F)
                cell = worksheet.cell(row=2, column=col_idx)
                cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border_style

            red_fill = PatternFill(fill_type="solid", fgColor="ffca2c")  # พื้นหลังสีแดง
            white_font = Font(color="FFFFFF")  # ตัวอักษรสีขาว

            # ปรับ Style สำหรับข้อมูลในตาราง
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if cell.value == "/":  # ถ้าเซลล์มี "/"
                        cell.fill = red_fill  # เปลี่ยนพื้นหลังเป็นสีแดง
                        # cell.font = white_font  # เปลี่ยนตัวอักษรเป็นสีขาว
                    cell.border = border_style
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # worksheet.freeze_panes = "F6"
            

        # ส่งไฟล์ Excel
        excel_file.seek(0)
        return excel_file