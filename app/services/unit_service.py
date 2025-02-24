from sqlalchemy.orm import Session
from sqlalchemy.sql import func,and_, or_
from sqlalchemy import desc
from app.models.unit import Unit
from app.models.position import Position
from app.models.dept import Dept
from app.models.mission import Mission
from app.models.mission_unit import MissionUnit


from app.schemas.unit import UnitCreate, UnitAll, UnitResponse, UnitUpdate

import pandas as pd
import os
import datetime
from datetime import timedelta

from io import BytesIO
from openpyxl.drawing.image import Image

from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def delete_unit(db: Session, unit_id: int):
    db.query(Unit).filter(Unit.units_id == unit_id).delete()

    # delete mission_unit
    db.query(MissionUnit).filter(MissionUnit.unit_id == unit_id).delete()

    db.commit()

    return {"message": "Delete success"}

def create_unit(db: Session, unit: UnitCreate):
    db_unit = Unit(**unit.dict())
    db.add(db_unit)
    db.commit()
    db.refresh(db_unit)
    return db_unit

def update_unit(db: Session, unit_id: int, unit: UnitUpdate):
    db.query(Unit).filter(Unit.units_id == unit_id).update(unit.dict())
    db.commit()
    return db.query(Unit).filter(Unit.units_id == unit_id).first()

def get_unit(db: Session, unit_id: int):
    query = db.query(
        Unit.units_id,
        Unit.first_name,
        Unit.last_name,
        Unit.tel,
        Unit.identify_id,
        Unit.position_id,
        Unit.dept_id,
        Unit.blood_group_id,
        Unit.province_id,
        Unit.district_id,
        Unit.parish_id,
        Unit.post_code,
        Unit.address_detail,
        Unit.status,
        Unit.is_active,
        Unit.img_path,
        Unit.identify_soldier_id,
        Position.position_name,
        Dept.dept_name,
        Unit.position_detail
    ).outerjoin(Position, Unit.position_id == Position.position_id) \
     .outerjoin(Dept, Unit.dept_id == Dept.dept_id) \
     .filter(Unit.units_id == unit_id) \
     .order_by(desc(Position.position_seq))\
     .first()

    if query:
        return {
            "units_id": query.units_id,
            "first_name": query.first_name,
            "last_name": query.last_name,
            "tel": query.tel,
            "identify_id": query.identify_id,
            "position_id": query.position_id,
            "dept_id": query.dept_id,
            "blood_group_id": query.blood_group_id,
            "province_id": query.province_id,
            "district_id": query.district_id,
            "parish_id": query.parish_id,
            "post_code": query.post_code,
            "address_detail": query.address_detail,
            "status": query.status,
            "is_active": query.is_active,
            "position_name": query.position_name,
            "dept_name": query.dept_name,
            "img_path": query.img_path,
            "identify_soldier_id": query.identify_soldier_id,
            "position_detail": query.position_detail
        }
    return None


def get_all_units(db: Session, name=None, position_id=None, dept_id=None,status=None,start_date = None, end_date= None, skip=0, limit=100):
    # Base query with join
    query = (
        db.query(
            Unit.units_id,
            Unit.first_name,
            Unit.last_name,
            Unit.img_path,
            Unit.status,
            Position.position_name,
            Position.position_seq,
            Dept.dept_name,
        )
        .outerjoin(Position, Unit.position_id == Position.position_id)
        .outerjoin(Dept, Unit.dept_id == Dept.dept_id)
    )

    # Apply filters
    if name:
        query = query.filter(Unit.first_name.contains(name) | Unit.last_name.contains(name))
    if position_id:
        query = query.filter(Unit.position_id == position_id)
    if dept_id:
        query = query.filter(Unit.dept_id == dept_id)
    if status:
        if status == 'ready':
            query = query.join(MissionUnit, MissionUnit.unit_id == Unit.units_id).\
            join(Mission, Mission.mission_id == MissionUnit.mission_id).\
            filter(MissionUnit.mission_id == Mission.mission_id).\
            filter(and_(
                Mission.mission_end < datetime.datetime.now(),
                Mission.mission_start < datetime.datetime.now(),
            ))
        elif status == 'not ready':
            query = query.join(MissionUnit, MissionUnit.unit_id == Unit.units_id).\
            join(Mission, Mission.mission_id == MissionUnit.mission_id).\
            filter(MissionUnit.mission_id == Mission.mission_id).\
            filter(Mission.mission_start <= datetime.datetime.now()).\
            filter(Mission.mission_end >= datetime.datetime.now())
        elif status == 'wait':
            query = query.join(MissionUnit, MissionUnit.unit_id == Unit.units_id).\
            join(Mission, Mission.mission_id == MissionUnit.mission_id).\
            filter(MissionUnit.mission_id == Mission.mission_id).\
            filter(Mission.mission_start > datetime.datetime.now())
        
    if start_date and end_date:
        mission = db.query(
            Mission.mission_id,
            MissionUnit.unit_id
        ).join(MissionUnit, MissionUnit.mission_id == Mission.mission_id).filter(
            or_(
                and_(Mission.mission_start >= start_date, Mission.mission_start <= end_date),
                and_(Mission.mission_end >= start_date, Mission.mission_end <= end_date),
                and_(Mission.mission_start <= start_date, Mission.mission_end >= end_date)  # ครอบคลุมกรณี mission ครอบทั้งช่วง
            )
        ).all()
        unit_id_list = [m.unit_id for m in mission]
        query = query.filter(Unit.units_id.not_in(unit_id_list))
    # Get total count for pagination
    # total = db.query(func.count(Unit.units_id)).scalar()
    total = query.with_entities(func.count()).scalar()

    if status:
        units = query.order_by(desc(Mission.mission_start),desc(Position.position_seq)).offset(skip).limit(limit).all()
        query = query.group_by(
                Unit.units_id,
                Unit.first_name,
                Unit.last_name,
                Unit.img_path,
                Unit.status,
                Position.position_name,
                Position.position_seq,
                Dept.dept_name)
    else:

        # Apply pagination
        units = query.order_by(desc(Position.position_seq)).offset(skip).limit(limit).all()

    data = []
    for unit in units:
        mission = db.query(
            MissionUnit.mission_id,
            MissionUnit.unit_id,
            Mission.mission_name,
            Mission.mission_start,
            Mission.mission_end,
        ).join(Mission, Mission.mission_id == MissionUnit.mission_id)\
        .filter(MissionUnit.unit_id == unit.units_id)\
        .order_by(desc(Mission.mission_start))\
        .first()



        status = None
        date_start = None
        date_end= None

        if mission:
            if isinstance(mission.mission_start, datetime.date):
                mission_start = datetime.datetime.combine(mission.mission_start, datetime.time.min)
            else:
                mission_start = mission.mission_start

            if isinstance(mission.mission_end, datetime.date):
                mission_end = datetime.datetime.combine(mission.mission_end, datetime.time.max)
            else:
                mission_end = mission.mission_end

            now = datetime.datetime.now()


            if mission_start > now:
                status = "รอเริ่มภารกิจ"
                date_start = mission.mission_start
                date_end = mission.mission_end
            elif mission_end < now:
                status = "ว่าง"
            else:
                status = "อยู่ในภารกิจ"
        else:
            status = "ว่าง"

        data.append({
            "units_id": unit.units_id,
            "first_name": unit.first_name,
            "last_name": unit.last_name,
            "position_name": unit.position_name,
            "dept_name": unit.dept_name,
            "status": status,
            "date_start": date_start,
            "date_end": date_end,
            "img_path": unit.img_path,
        })

    # Format the response
    return {
        "units": data,
        "total": total,
        "skip": skip,
        "limit": limit,
    }

def get_units_active(db, position_id, first_name,last_name, date_start, date_end):
    mission = db.query(
            Mission.mission_id,
            MissionUnit.unit_id
        ).join(MissionUnit, MissionUnit.mission_id == Mission.mission_id).\
        filter(
            or_(
                Mission.mission_end < date_start,  # Mission สิ้นสุดก่อนช่วง date_start
                Mission.mission_start > date_end  # Mission เริ่มหลังช่วง date_end
            )
        ).all()
    
    unit_id_list = [unit.unit_id for unit in mission]

    unit = db.query(
            Unit.units_id,
            Unit.first_name,
            Unit.last_name,
            Position.position_id,
            Position.position_name,
            Position.position_name_short,
            Position.position_seq,
        ).filter(
            or_(
                and_(
                Unit.status == 'ready',
                Unit.is_active == True
                ),
                Unit.units_id.in_(unit_id_list)
            )
        ).\
        join(Position, Position.position_id == Unit.position_id).order_by(desc(Position.position_seq))
        
    if first_name:
        unit = unit.filter(Unit.first_name.contains(first_name))
    if position_id:
        unit = unit.filter(Unit.position_id == position_id)
    if last_name:
        unit = unit.filter(Unit.last_name.contains(last_name))

    unit.all()

    return unit


def export_units(db):
    data = db.query(
                Position.position_name_short,
                Unit.first_name,
                Unit.last_name,
                Dept.dept_name_short,
                Unit.identify_id,
                Unit.identify_soldier_id,
                Unit.tel,
                Unit.blood_group_id,
                Unit.address_detail,
                Unit.position_detail
            ).\
            outerjoin(Position, Position.position_id == Unit.position_id).\
            outerjoin(Dept, Dept.dept_id == Unit.dept_id).\
            order_by(desc(Position.position_seq)).\
            all()
            

    orders_list = []
    index = 1
    for item in data:
        orders_list.append({
            "ลำดับ": index,
            "ยศ": item.position_name_short,
            "ชื่อ": item.first_name,
            "นามสกุล": item.last_name,
            "สังกัด": item.dept_name_short,
            "ตำแหน่ง": item.position_detail,
            "หมายเลขประจำตัว": item.identify_id,
            "หมายเลขประจำตัวทหาร": item.identify_soldier_id,
            "เบอร์โทร": item.tel,
            "กรุ๊ปเลือด": item.blood_group_id,
            "ที่อยู่": item.address_detail
        })
        index += 1
    # สร้าง DataFrame จากข้อมูล
    df = pd.DataFrame(orders_list)

    # สร้างไฟล์ Excel พร้อม Style
    excel_file = BytesIO()

    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:

        df.to_excel(writer, index=False, sheet_name="รายงาน")
        workbook = writer.book
        worksheet = writer.sheets["รายงาน"]

        # ปรับความกว้างของคอลัมน์
        for column_cells in worksheet.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 2
        
        # กำหนด Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        alignment = Alignment(horizontal="center", vertical="center")

        # จัดการหัวตาราง
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border_style
            cell.alignment = alignment

    # รีเซ็ต pointer และส่งไฟล์ 
    excel_file.seek(0)
    return excel_file

